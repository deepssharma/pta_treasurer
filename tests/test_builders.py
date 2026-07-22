"""
Tests for builders.py Excel sheet builder functions.
Uses mock data — no real files needed.
"""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

import pytest
import openpyxl
from builders import (build_treasurer, build_budget, build_givebacks,
                      build_manifest, build_credits_sheet, build_debits_sheet,
                      build_memberhub_summary_sheet, FISCAL_MONTHS)


# ── Mock data ─────────────────────────────────────────────────────────────────

MOCK_QRB = {
    'period':        'July 1-31, 2025',
    'income':        {},
    'income_total':  0.0,
    'expenses':      {'Accounting Quickbooks': 1257.76, '6th Grade Events': 181.58},
    'expense_total': 1439.34,
    'net_income':    -1439.34,
    'transactions':  [],
}

MOCK_BANK = {
    'period':            'July 01, 2025 through July 31, 2025',
    'account':           '4346',
    'beginning_balance': 32630.10,
    'ending_balance':    31190.76,
    'total_deposits':    0.0,
    'total_checks':      181.58,
    'total_withdrawals': 1257.76,
    'total_fees':        0.0,
    'deposits':          [],
    'checks':            [{'check_no': '1077', 'date': '07/01', 'amount': 181.58}],
    'withdrawals':       [],
    'fees':              [],
    'daily_balances':    {'07/31': 31190.76},
    'source_file':       'Chase_july_2025.pdf',
}

MOCK_GIVEBACKS = [
    {'item': 'Shop to Give Donation', 'category': '',
     'count': 1, 'total': 95.14, 'source_file': 'givebacks_july.csv'},
    {'item': 'Teacher/Staff', 'category': 'Memberships',
     'count': 1, 'total': 15.0,  'source_file': 'givebacks_july.csv'},
]


# ── Treasurer Report tests ────────────────────────────────────────────────────

def test_build_treasurer_creates_sheet():
    wb = openpyxl.Workbook()
    ws = wb.active
    build_treasurer(ws, MOCK_QRB, MOCK_BANK, 'July 2025', 'Setauket School PTA')
    assert ws['A1'].value == 'SETAUKET SCHOOL PTA'

def test_build_treasurer_has_org_name():
    wb = openpyxl.Workbook()
    ws = wb.active
    build_treasurer(ws, MOCK_QRB, MOCK_BANK, 'July 2025', 'Test PTA')
    assert ws['A1'].value == 'TEST PTA'

def test_build_treasurer_bank_reconciliation():
    wb = openpyxl.Workbook()
    ws = wb.active
    build_treasurer(ws, MOCK_QRB, MOCK_BANK, 'July 2025', 'Setauket School PTA')
    # Find beginning balance cell
    found = False
    for row in ws.iter_rows(values_only=True):
        if row[0] == 'Beginning Balance':
            assert row[1] == 32630.10
            found = True
            break
    assert found, 'Beginning Balance row not found'

def test_build_treasurer_difference_is_zero():
    wb = openpyxl.Workbook()
    ws = wb.active
    build_treasurer(ws, MOCK_QRB, MOCK_BANK, 'July 2025', 'Setauket School PTA')
    found = False
    for row in ws.iter_rows(values_only=True):
        if row[0] == 'Difference (should be $0.00)':
            assert abs(row[1]) < 0.01
            found = True
            break
    assert found, 'Difference row not found'


# ── Budget sheet tests ────────────────────────────────────────────────────────

def test_build_budget_income(sample_merged_income):
    wb = openpyxl.Workbook()
    ws = wb.active
    build_budget(ws, 'Budget vs Actuals - Income',
                 sample_merged_income, 'Setauket School PTA',
                 FISCAL_MONTHS, 0, show_pl=False)
    assert 'SETAUKET SCHOOL PTA' in ws['A1'].value

def test_build_budget_active_month_highlighted(sample_merged_income):
    wb = openpyxl.Workbook()
    ws = wb.active
    build_budget(ws, 'Test', sample_merged_income,
                 'Test PTA', FISCAL_MONTHS, 0, show_pl=False)
    # Column D (index 4) = JULY = fiscal index 0 — should be gold
    header_cell = ws.cell(row=3, column=4)
    # openpyxl stores colors as 8-char ARGB (alpha + RGB)
    # so FFD966 becomes 00FFD966
    assert header_cell.fill.fgColor.rgb.endswith('FFD966')  # ← use endswith

def test_build_budget_expense_with_pl(sample_merged_expense, sample_merged_income):
    wb = openpyxl.Workbook()
    ws = wb.active
    build_budget(ws, 'Budget vs Actuals - Expenses',
                 sample_merged_expense, 'Test PTA',
                 FISCAL_MONTHS, 0, show_pl=True,
                 income_merged=sample_merged_income)
    # Should have Profit/Loss header
    found = False
    for row in ws.iter_rows(min_row=3, max_row=3, values_only=True):
        if 'Profit/Loss' in [v for v in row if v]:
            found = True
    assert found, 'Profit/Loss column not found'


# ── Givebacks reconciliation tests ───────────────────────────────────────────

def test_build_givebacks_total():
    wb = openpyxl.Workbook()
    ws = wb.active
    build_givebacks(ws, MOCK_GIVEBACKS, MOCK_BANK, 'Test PTA')
    # Check total row exists
    found = False
    for row in ws.iter_rows(values_only=True):
        if row[0] == 'TOTAL':
            found = True
            break
    assert found, 'TOTAL row not found'

def test_build_givebacks_item_count():
    wb = openpyxl.Workbook()
    ws = wb.active
    build_givebacks(ws, MOCK_GIVEBACKS, MOCK_BANK, 'Test PTA')
    items_found = 0
    for row in ws.iter_rows(values_only=True):
        if row[0] in ('Shop to Give Donation', 'Teacher/Staff'):
            items_found += 1
    assert items_found == 2


# ── Fiscal months constants test ──────────────────────────────────────────────

def test_fiscal_months_count():
    assert len(FISCAL_MONTHS) == 12

def test_fiscal_months_order():
    assert FISCAL_MONTHS[0]  == 'JULY'
    assert FISCAL_MONTHS[5]  == 'DEC'
    assert FISCAL_MONTHS[6]  == 'JAN'
    assert FISCAL_MONTHS[11] == 'JUNE'


# ── Debits & Credits (whole-fiscal-year ledger) tests ────────────────────────

MOCK_CREDITS_BY_MONTH = [
    ('July 2025', [
        {'date': '07/17/2025', 'type': 'Deposit', 'check_no': '', 'payee': 'MemberHub',
         'description': 'Deposit', 'category': 'Membership', 'amount': 110.14, 'is_income': True},
    ]),
    ('August 2025', [
        {'date': '08/05/2025', 'type': 'Deposit', 'check_no': '', 'payee': 'MemberHub',
         'description': 'Deposit', 'category': 'Book Fair', 'amount': 50.0, 'is_income': True},
    ]),
    ('September 2025', []),  # no credits that month - should be skipped, not error
]

MOCK_DEBITS_BY_MONTH = [
    ('July 2025', [
        {'date': '07/01/2025', 'type': 'Check', 'check_no': '1077', 'payee': 'Jane Doe',
         'description': 'CHECK # 1077', 'category': 'Picnic', 'amount': 181.58, 'is_income': False},
        {'date': '07/17/2025', 'type': 'Expense', 'check_no': '', 'payee': 'Quickbooks Online',
         'description': 'Accounting', 'category': 'Accounting Expense (Quickbooks)',
         'amount': 1257.76, 'is_income': False},
    ]),
]

MOCK_GIVEBACKS_BY_MONTH = [
    ('July 2025', [
        {'item': 'Shop to Give Donation', 'category': '', 'count': 1, 'total': 95.14, 'source_file': 'x.csv'},
        {'item': 'Teacher/Staff', 'category': 'Memberships', 'count': 1, 'total': 15.0, 'source_file': 'x.csv'},
    ]),
]

MOCK_QB_TO_BUDGET_MAP = {'Membership': 'Membership Income', 'Picnic': 'Picnic Fund'}


def _cell_values(ws):
    return [row for row in ws.iter_rows(values_only=True) if any(v is not None for v in row)]


def test_build_credits_sheet_running_total():
    wb = openpyxl.Workbook()
    ws = wb.active
    build_credits_sheet(ws, MOCK_CREDITS_BY_MONTH, 'Test PTA', MOCK_QB_TO_BUDGET_MAP)
    rows = _cell_values(ws)
    amounts = [r[5] for r in rows if isinstance(r[0], str) and '/' in str(r[0])]
    assert amounts == [110.14, 160.14]  # cumulative across months


def test_build_credits_sheet_skips_empty_month():
    wb = openpyxl.Workbook()
    ws = wb.active
    build_credits_sheet(ws, MOCK_CREDITS_BY_MONTH, 'Test PTA')
    rows = _cell_values(ws)
    month_bands = [r[0] for r in rows if r[0] in ('JULY', 'AUGUST', 'SEPTEMBER')]
    assert month_bands == ['JULY', 'AUGUST']  # September (empty) skipped


def test_build_credits_sheet_budget_line_mapping():
    wb = openpyxl.Workbook()
    ws = wb.active
    build_credits_sheet(ws, MOCK_CREDITS_BY_MONTH, 'Test PTA', MOCK_QB_TO_BUDGET_MAP)
    rows = _cell_values(ws)
    budget_lines = [r[4] for r in rows if isinstance(r[0], str) and '/' in str(r[0])]
    assert budget_lines == ['Membership Income', 'Book Fair']  # mapped / falls back to raw category


def test_build_credits_sheet_total_row():
    wb = openpyxl.Workbook()
    ws = wb.active
    build_credits_sheet(ws, MOCK_CREDITS_BY_MONTH, 'Test PTA')
    rows = _cell_values(ws)
    total_rows = [r for r in rows if r[0] == 'TOTAL CREDITS']
    assert len(total_rows) == 1
    assert total_rows[0][2] == 160.14


def test_build_debits_sheet_running_total_and_notes_blank():
    wb = openpyxl.Workbook()
    ws = wb.active
    build_debits_sheet(ws, MOCK_DEBITS_BY_MONTH, 'Test PTA', MOCK_QB_TO_BUDGET_MAP)
    rows = _cell_values(ws)
    data_rows = [r for r in rows if r[0] in ('1077', '')]
    assert len(data_rows) == 2
    assert data_rows[1][8] == 1439.34  # running total after both debits
    for r in data_rows:
        assert not r[7]  # NOTES column intentionally blank


def test_build_debits_sheet_total_row():
    wb = openpyxl.Workbook()
    ws = wb.active
    build_debits_sheet(ws, MOCK_DEBITS_BY_MONTH, 'Test PTA')
    rows = _cell_values(ws)
    total_rows = [r for r in rows if r[0] == 'TOTAL DEBITS']
    assert len(total_rows) == 1
    assert total_rows[0][4] == 1439.34


def test_build_memberhub_summary_statement_total_once_per_month():
    wb = openpyxl.Workbook()
    ws = wb.active
    build_memberhub_summary_sheet(ws, MOCK_GIVEBACKS_BY_MONTH, 'Test PTA')
    rows = _cell_values(ws)
    item_rows = [r for r in rows if r[0] in ('Shop to Give Donation', 'Teacher/Staff')]
    assert item_rows[0][3] == 110.14   # statement total on first item row
    assert item_rows[1][3] is None     # not repeated on subsequent rows


def test_build_memberhub_summary_running_total():
    wb = openpyxl.Workbook()
    ws = wb.active
    build_memberhub_summary_sheet(ws, MOCK_GIVEBACKS_BY_MONTH, 'Test PTA')
    rows = _cell_values(ws)
    item_rows = [r for r in rows if r[0] in ('Shop to Give Donation', 'Teacher/Staff')]
    assert [r[4] for r in item_rows] == [95.14, 110.14]
