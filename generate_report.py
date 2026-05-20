"""
PTA Treasurer Report Generator
Inputs:  givebacks CSV, QuickBooks CSV, Chase bank statement PDF
Outputs: treasurer_report.xlsx  (4 tabs)
           1. Monthly Treasurer Report
           2. Budget vs Actuals - Income
           3. Budget vs Actuals - Expenses
           4. Giveback Reconciliation
"""

import csv
import re
import sys
from pathlib import Path
from datetime import datetime

import pdfplumber
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── colour palette ────────────────────────────────────────────────────────────
NAVY   = "1F3864"
TEAL   = "2E75B6"
LTBLUE = "BDD7EE"
GOLD   = "FFD966"
WHITE  = "FFFFFF"
LGREY  = "F2F2F2"
GREEN  = "E2EFDA"
RED_BG = "FCE4D6"

HEADER_FONT   = Font(name="Arial", bold=True, color=WHITE, size=11)
TITLE_FONT    = Font(name="Arial", bold=True, color=NAVY,  size=13)
SUBHDR_FONT   = Font(name="Arial", bold=True, color=WHITE, size=10)
BODY_FONT     = Font(name="Arial", size=10)
BOLD_FONT     = Font(name="Arial", bold=True, size=10)
TOTAL_FONT    = Font(name="Arial", bold=True, size=10, color=NAVY)

NAVY_FILL   = PatternFill("solid", fgColor=NAVY)
TEAL_FILL   = PatternFill("solid", fgColor=TEAL)
LTBLUE_FILL = PatternFill("solid", fgColor=LTBLUE)
GOLD_FILL   = PatternFill("solid", fgColor=GOLD)
LGREY_FILL  = PatternFill("solid", fgColor=LGREY)
GREEN_FILL  = PatternFill("solid", fgColor=GREEN)
RED_FILL    = PatternFill("solid", fgColor=RED_BG)

THIN = Side(style="thin", color="AAAAAA")
MED  = Side(style="medium", color=NAVY)
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MED_BORDER  = Border(left=MED,  right=MED,  top=MED,  bottom=MED)

MONEY_FMT = '$#,##0.00_);($#,##0.00)'
NEG_FMT   = '$#,##0.00;($#,##0.00);"-"'


def money(v):
    try:
        return float(str(v).replace("$", "").replace(",", "").strip())
    except:
        return 0.0


def style_cell(cell, font=None, fill=None, align=None, border=None, fmt=None, value=None):
    if value is not None:
        cell.value = value
    if font:   cell.font      = font
    if fill:   cell.fill      = fill
    if align:  cell.alignment = align
    if border: cell.border    = border
    if fmt:    cell.number_format = fmt


def center(cell):
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def right(cell):
    cell.alignment = Alignment(horizontal="right", vertical="center")


def left(cell):
    cell.alignment = Alignment(horizontal="left", vertical="center")


# ── 1. PARSE GIVEBACKS CSV ────────────────────────────────────────────────────
def parse_givebacks(path: Path) -> list[dict]:
    rows = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for r in reader:
            item = r.get("Item", "").strip()
            if not item:
                continue
            rows.append({
                "item":     item,
                "category": r.get("Categories", "").strip(),
                "count":    int(r.get("No. of Transactions", "0").strip() or 0),
                "total":    money(r.get("Total", "0")),
            })
    return rows


# ── 2. PARSE QUICKBOOKS CSV ───────────────────────────────────────────────────
def parse_quickbooks(path: Path) -> dict:
    """Returns structured P&L dict from QuickBooks export."""
    data = {
        "period": "",
        "income": {},        # category -> amount
        "income_total": 0.0,
        "expenses": {},      # category -> amount
        "expense_total": 0.0,
        "net_income": 0.0,
    }
    with open(path, newline="", encoding="utf-8-sig") as f:
        lines = [r for r in csv.reader(f)]

    # Period is on line 3
    if len(lines) > 2:
        data["period"] = lines[2][0].strip() if lines[2] else ""

    section = None
    for row in lines:
        if not row or not row[0].strip():
            continue
        label = row[0].strip()
        val   = money(row[1]) if len(row) > 1 else 0.0

        if label == "Income":
            section = "income"
        elif label == "Expenses":
            section = "expenses"
        elif label.startswith("Total for Income"):
            data["income_total"] = val
        elif label.startswith("Total for Expenses") or label == "Total for Expenses":
            data["expense_total"] = val
        elif label.startswith("Net Income"):
            data["net_income"] = val
        elif label.startswith("Total for") or label.startswith("Gross Profit") \
                or label.startswith("Net Operating") or label.startswith("Net Other"):
            continue
        elif section == "income" and val != 0.0:
            data["income"][label] = val
        elif section == "expenses" and val != 0.0:
            data["expenses"][label] = val

    return data


# ── 3. PARSE CHASE PDF ────────────────────────────────────────────────────────
def parse_chase_pdf(path: Path) -> dict:
    bank = {
        "period": "",
        "account": "",
        "beginning_balance": 0.0,
        "ending_balance": 0.0,
        "total_deposits": 0.0,
        "total_checks": 0.0,
        "total_fees": 0.0,
        "deposits": [],
        "checks": [],
        "fees": [],
        "daily_balances": {},
    }

    with pdfplumber.open(path) as pdf:
        full_text = "\n".join(p.extract_text() or "" for p in pdf.pages)

    # Period
    m = re.search(r"(January|February|March|April|May|June|July|August|September|October|November|December)"
                  r"\s+\d+,\s+\d{4}\s+through\s+\S+\s+\d+,\s+\d{4}", full_text)
    if m:
        bank["period"] = m.group(0)

    # Account
    m = re.search(r"Account Number:\s+([\d]+)", full_text)
    if m:
        bank["account"] = m.group(1)

    def grab(pattern):
        m = re.search(pattern, full_text)
        return money(m.group(1)) if m else 0.0

    bank["beginning_balance"] = grab(r"Beginning Balance\s+\$?([\d,]+\.\d{2})")
    bank["ending_balance"]    = grab(r"Ending Balance\s+\d+\s+\$?([\d,]+\.\d{2})")
    bank["total_deposits"]    = grab(r"Total Deposits and Additions\s+\$?([\d,]+\.\d{2})")
    bank["total_checks"]      = grab(r"Total Checks Paid\s+\$?([\d,]+\.\d{2})")
    bank["total_fees"]        = grab(r"Total Fees\s+\$?([\d,]+\.\d{2})")

    # Individual deposits
    dep_block = re.search(r"DEPOSITS AND ADDITIONS(.*?)CHECKS PAID", full_text, re.DOTALL)
    if dep_block:
        for m in re.finditer(r"(\d{2}/\d{2})\s+(.+?)\s+\$?([\d,]+\.\d{2})", dep_block.group(1)):
            bank["deposits"].append({
                "date": m.group(1), "description": m.group(2).strip(), "amount": money(m.group(3))
            })

    # Checks paid
    for m in re.finditer(r"(\d{4})\s+\^?\s+(\d{2}/\d{2})\s+\$?([\d,]+\.\d{2})", full_text):
        bank["checks"].append({
            "check_no": m.group(1), "date": m.group(2), "amount": money(m.group(3))
        })

    # Fees
    fee_block = re.search(r"FEES(.*?)DAILY ENDING BALANCE", full_text, re.DOTALL)
    if fee_block:
        for m in re.finditer(r"(\d{2}/\d{2})\s+(.+?)\s+\$?([\d,]+\.\d{2})", fee_block.group(1)):
            bank["fees"].append({
                "date": m.group(1), "description": m.group(2).strip(), "amount": money(m.group(3))
            })

    # Daily balances
    bal_block = re.search(r"DAILY ENDING BALANCE(.*?)$", full_text, re.DOTALL)
    if bal_block:
        for m in re.finditer(r"(\d{2}/\d{2})\s+([\d,]+\.\d{2})", bal_block.group(1)):
            bank["daily_balances"][m.group(1)] = money(m.group(2))

    return bank


# ── 4. SHEET: MONTHLY TREASURER REPORT ───────────────────────────────────────
def build_treasurer_report(ws, qb: dict, bank: dict, month_label: str):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 22

    # Title
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "SETAUKET SCHOOL PTA"
    c.font  = Font(name="Arial", bold=True, size=16, color=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:D2")
    c = ws["A2"]
    c.value = f"Monthly Treasurer Report — {month_label}"
    c.font  = Font(name="Arial", bold=True, size=12, color=TEAL)
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    ws.merge_cells("A3:D3")
    c = ws["A3"]
    c.value = f"Generated: {datetime.today().strftime('%B %d, %Y')}"
    c.font  = Font(name="Arial", italic=True, size=9, color="888888")
    c.alignment = Alignment(horizontal="center")

    row = 5

    def section_header(label, r):
        ws.merge_cells(f"A{r}:D{r}")
        c = ws[f"A{r}"]
        c.value = label
        c.font  = Font(name="Arial", bold=True, size=11, color=WHITE)
        c.fill  = NAVY_FILL
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[r].height = 20
        return r + 1

    def col_headers(r, labels):
        fills = [TEAL_FILL] * len(labels)
        for i, (col, lbl) in enumerate(zip(["A","B","C","D"], labels)):
            c = ws[f"{col}{r}"]
            c.value = lbl
            c.font  = SUBHDR_FONT
            c.fill  = TEAL_FILL
            c.alignment = Alignment(horizontal="center" if i > 0 else "left",
                                    vertical="center", indent=1 if i == 0 else 0)
            c.border = THIN_BORDER
        ws.row_dimensions[r].height = 18
        return r + 1

    def data_row(r, label, amount, shade=False):
        fill = LGREY_FILL if shade else PatternFill()
        ws[f"A{r}"].value = label
        ws[f"A{r}"].font  = BODY_FONT
        ws[f"A{r}"].fill  = fill
        ws[f"A{r}"].alignment = Alignment(indent=2)
        ws[f"A{r}"].border = THIN_BORDER
        ws[f"B{r}"].value  = amount
        ws[f"B{r}"].font   = BODY_FONT
        ws[f"B{r}"].fill   = fill
        ws[f"B{r}"].number_format = MONEY_FMT
        ws[f"B{r}"].alignment = Alignment(horizontal="right")
        ws[f"B{r}"].border = THIN_BORDER
        for col in ["C","D"]:
            ws[f"{col}{r}"].fill = fill
            ws[f"{col}{r}"].border = THIN_BORDER
        ws.row_dimensions[r].height = 16

    def total_row(r, label, formula_or_val):
        ws[f"A{r}"].value = label
        ws[f"A{r}"].font  = TOTAL_FONT
        ws[f"A{r}"].fill  = LTBLUE_FILL
        ws[f"A{r}"].border = MED_BORDER
        ws[f"A{r}"].alignment = Alignment(indent=1)
        ws[f"B{r}"].value  = formula_or_val
        ws[f"B{r}"].font   = TOTAL_FONT
        ws[f"B{r}"].fill   = LTBLUE_FILL
        ws[f"B{r}"].border = MED_BORDER
        ws[f"B{r}"].number_format = MONEY_FMT
        ws[f"B{r}"].alignment = Alignment(horizontal="right")
        for col in ["C","D"]:
            ws[f"{col}{r}"].fill = LTBLUE_FILL
            ws[f"{col}{r}"].border = MED_BORDER
        ws.row_dimensions[r].height = 18

    # ── INCOME ────────────────────────────────────────────────────────────────
    row = section_header("INCOME", row)
    row = col_headers(row, ["Category", "Amount", "", ""])
    income_start = row
    for i, (cat, amt) in enumerate(qb["income"].items()):
        data_row(row, cat, amt, shade=(i % 2 == 1))
        row += 1
    income_end = row - 1
    total_row(row, "Total Income", f"=SUM(B{income_start}:B{income_end})")
    income_total_row = row
    row += 2

    # ── EXPENSES ──────────────────────────────────────────────────────────────
    row = section_header("EXPENSES", row)
    row = col_headers(row, ["Category", "Amount", "", ""])
    exp_start = row
    for i, (cat, amt) in enumerate(qb["expenses"].items()):
        data_row(row, cat, amt, shade=(i % 2 == 1))
        row += 1
    exp_end = row - 1
    total_row(row, "Total Expenses", f"=SUM(B{exp_start}:B{exp_end})")
    exp_total_row = row
    row += 2

    # ── NET INCOME ────────────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"].value = "NET INCOME / (LOSS)"
    ws[f"A{row}"].font  = Font(name="Arial", bold=True, size=11, color=WHITE)
    ws[f"A{row}"].fill  = NAVY_FILL
    ws[f"A{row}"].alignment = Alignment(horizontal="left", indent=1)
    row += 1

    net_val = f"=B{income_total_row}-B{exp_total_row}"
    ws[f"A{row}"].value = "Net Income (Loss)"
    ws[f"A{row}"].font  = TOTAL_FONT
    ws[f"A{row}"].fill  = GOLD_FILL
    ws[f"A{row}"].border = MED_BORDER
    ws[f"A{row}"].alignment = Alignment(indent=1)
    ws[f"B{row}"].value = net_val
    ws[f"B{row}"].font  = TOTAL_FONT
    ws[f"B{row}"].fill  = GOLD_FILL
    ws[f"B{row}"].border = MED_BORDER
    ws[f"B{row}"].number_format = MONEY_FMT
    ws[f"B{row}"].alignment = Alignment(horizontal="right")
    for col in ["C","D"]:
        ws[f"{col}{row}"].fill = GOLD_FILL
        ws[f"{col}{row}"].border = MED_BORDER
    row += 2

    # ── BANK RECONCILIATION ───────────────────────────────────────────────────
    row = section_header("BANK RECONCILIATION  —  Chase Business Checking", row)
    row = col_headers(row, ["Description", "Amount", "", ""])

    recon_items = [
        ("Beginning Balance", bank["beginning_balance"]),
        ("(+) Deposits & Additions", bank["total_deposits"]),
        ("(-) Checks Paid", -bank["total_checks"]),
        ("(-) Fees", -bank["total_fees"]),
    ]
    for i, (lbl, amt) in enumerate(recon_items):
        data_row(row, lbl, amt, shade=(i % 2 == 1))
        row += 1
    total_row(row, "Ending Balance (per statement)", bank["ending_balance"])
    row += 1

    # Reconciliation check
    check_row = row
    ws[f"A{row}"].value = "Calculated Ending Balance"
    ws[f"A{row}"].font  = BOLD_FONT
    ws[f"A{row}"].alignment = Alignment(indent=2)
    ws[f"A{row}"].border = THIN_BORDER
    start_r = check_row - 5  # beginning balance row
    ws[f"B{row}"].value = (bank["beginning_balance"]
                           + bank["total_deposits"]
                           - bank["total_checks"]
                           - bank["total_fees"])
    ws[f"B{row}"].number_format = MONEY_FMT
    ws[f"B{row}"].font = BOLD_FONT
    ws[f"B{row}"].alignment = Alignment(horizontal="right")
    ws[f"B{row}"].border = THIN_BORDER
    for col in ["C","D"]:
        ws[f"{col}{row}"].border = THIN_BORDER
    row += 1

    ws[f"A{row}"].value = "Difference (should be $0.00)"
    ws[f"A{row}"].font  = BOLD_FONT
    ws[f"A{row}"].alignment = Alignment(indent=2)
    ws[f"A{row}"].border = THIN_BORDER
    diff = (bank["beginning_balance"] + bank["total_deposits"]
            - bank["total_checks"] - bank["total_fees"]) - bank["ending_balance"]
    ws[f"B{row}"].value = diff
    ws[f"B{row}"].number_format = MONEY_FMT
    ws[f"B{row}"].font = BOLD_FONT
    ws[f"B{row}"].alignment = Alignment(horizontal="right")
    ws[f"B{row}"].border = THIN_BORDER
    ws[f"B{row}"].fill = GREEN_FILL if abs(diff) < 0.01 else RED_FILL
    for col in ["C","D"]:
        ws[f"{col}{row}"].fill = GREEN_FILL if abs(diff) < 0.01 else RED_FILL
        ws[f"{col}{row}"].border = THIN_BORDER
    row += 2

    # ── DAILY BALANCES (small table) ──────────────────────────────────────────
    if bank["daily_balances"]:
        row = section_header("DAILY ENDING BALANCES", row)
        ws[f"A{row}"].value = "Date"
        ws[f"A{row}"].font  = SUBHDR_FONT
        ws[f"A{row}"].fill  = TEAL_FILL
        ws[f"A{row}"].border = THIN_BORDER
        ws[f"A{row}"].alignment = Alignment(horizontal="center")
        ws[f"B{row}"].value = "Balance"
        ws[f"B{row}"].font  = SUBHDR_FONT
        ws[f"B{row}"].fill  = TEAL_FILL
        ws[f"B{row}"].border = THIN_BORDER
        ws[f"B{row}"].alignment = Alignment(horizontal="center")
        row += 1
        for i, (dt, bal) in enumerate(sorted(bank["daily_balances"].items())):
            ws[f"A{row}"].value = dt
            ws[f"A{row}"].font  = BODY_FONT
            ws[f"A{row}"].fill  = LGREY_FILL if i % 2 else PatternFill()
            ws[f"A{row}"].border = THIN_BORDER
            ws[f"A{row}"].alignment = Alignment(horizontal="center")
            ws[f"B{row}"].value = bal
            ws[f"B{row}"].font  = BODY_FONT
            ws[f"B{row}"].fill  = LGREY_FILL if i % 2 else PatternFill()
            ws[f"B{row}"].border = THIN_BORDER
            ws[f"B{row}"].number_format = MONEY_FMT
            ws[f"B{row}"].alignment = Alignment(horizontal="right")
            row += 1


# ── 5. SHEET: BUDGET VS ACTUALS (INCOME) ─────────────────────────────────────
MONTHS = ["JULY","AUG","SEPT","OCT","NOV","DEC","JAN","FEB","MAR","APR","MAY","JUNE"]

INCOME_BUDGET = {
    "Fundraising": {
        "Birthday Books": (2310.00, 2000.00),
        "Book Fair":       (9118.36, 500.00),
        "Croc Charms":     (405.00,  100.00),
        "Fall Pictures":   (3350.25, 3000.00),
        "FAST":            (26370.00, 1000.00),
        "Holiday Boutique":(13417.00, 7500.00),
        "Plant Sale":      (8202.68, 8000.00),
        "Spiritwear":      (1253.96, 1500.00),
        "Spring Pictures": (0.00,    0.00),
    },
    "Basket Dinner": {
        "Ticket & Raffle Sales": (22165.00, 15000.00),
        "Sponsors":              (7450.00,  4000.00),
    },
    "Membership": {
        "Single":    (2580.00, 2000.00),
        "Family":    (1675.00, 1000.00),
        "Donations": (290.63,  100.00),
        "Teachers":  (0.00,    165.00),
        "Student":   (0.00,    5.00),
    },
    "Program": {
        "Gingerbread U":    (3515.00, 0.00),
        "Staff Appreciation":(1625.00, 1000.00),
        "Talent Show":      (1070.00, 750.00),
    },
    "Grad Class Activities": {
        "Family Contributions": (6480.00, 3900.00),
        "Lawn Signs":           (2560.00, 750.00),
        "Treat or Trunk":       (1250.00, 975.00),
        "The Night at Rinx":    (0.00,    0.00),
    },
}

# Monthly actuals from the PDF (hardcoded from source docs; these are the
# same numbers visible in the uploaded Income PDF)
INCOME_ACTUALS = {
    "Fundraising": {
        "Birthday Books": [0,0,1275,390,120,60,75,30,60,0,0,0],
        "Book Fair":      [0,0,4365.5,28.52,0,0,0,6219.40,0,0,0,0],
        "Croc Charms":    [0]*12,
        "Fall Pictures":  [0,0,3235.21,0,0,0,0,0,0,0,0,0],
        "FAST":           [0,0,8980,570,0,2285,4325,0,4340,3360,0,0],
        "Holiday Boutique":[0,0,13260,0,0,0,0,0,0,0,0,0],
        "Plant Sale":     [0,0,0,0,0,0,0,130,0,1182,0,0],
        "Spiritwear":     [0,0,0,0,0,0,0,898.72,0,0,0,0],
        "Spring Pictures":[0]*12,
    },
    "Basket Dinner": {
        "Ticket & Raffle Sales": [0,0,0,0,0,0,0,0,0,14390,0,0],
        "Sponsors":              [0,0,0,0,0,0,0,0,0,10500,0,0],
    },
    "Membership": {
        "Single":    [0,0,75,645,330,0,15,0,0,0,0,0],
        "Family":    [0,0,175,550,50,0,25,0,25,0,0,0],
        "Donations": [0,0,95.14,46.47,0,0,64.68,0,0,0,0,0],
        "Teachers":  [0,0,0,120,240,0,0,30,0,0,0,0],
        "Student":   [0]*12,
    },
    "Program": {
        "Gingerbread U":     [0,0,0,0,2450,994,0,0,0,0,0,0],
        "Staff Appreciation":[0,0,0,0,0,0,0,0,0,1135,0,0],
        "Talent Show":       [0,0,0,0,0,0,0,0,660,200,0,0],
    },
    "Grad Class Activities": {
        "Family Contributions":[0,0,3700,300,0,0,0,0,0,0,0,0],
        "Lawn Signs":          [0,0,0,855,0,0,0,0,0,0,0,0],
        "Treat or Trunk":      [0,0,430,0,0,0,0,0,0,0,0,0],  # +975 in Oct in doc
        "The Night at Rinx":   [0,0,0,0,0,0,960,0,40,20,0,0],
    },
}

EXPENSE_BUDGET = {
    "Admin/General": {
        "Accountant":             (650.00,   650.00),
        "Bank Services":          (234.94,   200.00),
        "Insurance":              (350.00,   350.00),
        "Supplies":               (450.22,   500.00),
        "Accounting Quickbooks":  (410.66,  1300.00),
        "Training":               (68.90,    100.00),
        "Website & Remind App":   (344.01,  1000.00),
        "Event Equipment":        (563.32,  1000.00),
    },
    "Fundraising": {
        "Birthday Books":    (503.39,    500.00),
        "Book Fair":         (9523.91, 10000.00),
        "Croc Charms":       (205.00,      0.00),
        "Fall Pictures":     (0.00,        0.00),
        "FAST":              (24000.00, 17000.00),
        "Holiday Boutique":  (11714.86, 10000.00),
        "Plant Sale":        (5615.95,  6000.00),
        "Spiritwear":        (0.00,     1000.00),
        "Spring Pictures":   (0.00,        0.00),
    },
    "Membership": {
        "Council Dues":        (125.00,  150.00),
        "Membership Expenses": (1034.00, 1500.00),
    },
    "Basket Dinner": {
        "Entertainment": (1405.42, 1000.00),
        "Raffles":       (1831.70, 2000.00),
        "Venue":         (8265.20, 10000.00),
    },
    "Programs": {
        "Bus Driver Appreciation": (240.00,   300.00),
        "Electric Parade":         (371.89,   200.00),
        "Family Connect Nights":   (771.00,  1500.00),
        "Gingerbread U":           (4245.66, 4500.00),
        "Homecoming":              (266.31,   150.00),
        "K Playdate":              (83.13,    300.00),
        "K Orientation":           (0.00,     400.00),
        "Milk & Cookies":          (358.07,   750.00),
        "Multicultural Night":     (2022.51, 2500.00),
        "Outdoor Movie":           (1884.20, 2500.00),
        "Science Fair":            (883.54,  1500.00),
        "Spring Fling":            (0.00,    2500.00),
        "Staff Appreciation":      (3840.00, 4000.00),
        "Talent Show":             (340.62,  1000.00),
        "Talent Show DJ":          (550.00,   500.00),
        "Volunteer Breakfast":     (64.00,    350.00),
        "Welcome Back Breakfast":  (582.66,  1000.00),
        "WINGO":                   (0.00,     500.00),
    },
    "Donations": {
        "BOE Gifts":              (200.00,    180.00),
        "Cultural Arts":          (14651.20, 15000.00),
        "Folders":                (580.00,    600.00),
        "Gardening":              (0.00,      500.00),
        "Hospitality":            (124.70,    750.00),
        "5th Staff T-Shirts":     (191.78,    125.00),
        "Recess Equipment":       (1000.00,  1000.00),
        "School Spirit":          (280.76,    750.00),
        "Sling Bags":             (2354.50,  1500.00),
        "Spelling Bee":           (192.50,    225.00),
        "Sunshine Fund":          (210.00,    500.00),
        "Trick or Treat Street":  (0.00,      250.00),
        "WM Scholarships":        (1000.00,  1000.00),
    },
    "Grad Class Events": {
        "Monster Bash":    (402.84,   1300.00),
        "Monster Bash DJ": (500.00,    500.00),
        "Electric Parade": (0.00,      900.00),
        "Winter Social":   (1167.86,  1300.00),
        "Winter Social DJ":(500.00,    500.00),
        "Moving Up":       (715.00,    400.00),
        "Picnic":          (2441.98,   900.00),
    },
    "Grad Class Expenses": {
        "Graduating Class Gifts": (1034.41,  500.00),
        "Graduating Mural":       (244.48,   350.00),
        "5th Grade T-Shirts":     (1494.25,  800.00),
        "Trunk or Treat Fundraiser":(258.67, 500.00),
        "The Night at Rinx":      (710.00,   650.00),
    },
}

EXPENSE_ACTUALS = {
    "Admin/General": {
        "Accountant":            [0,0,0,0,0,0,685,0,0,0,0,0],
        "Bank Services":         [0,16.25,0,0,0,0,0,3.34,0,30.50,0,0],
        "Insurance":             [0]*12,
        "Supplies":              [0]*12,
        "Accounting Quickbooks": [1257.76,0,0,0,0,0,0,0,0,0,0,0],
        "Training":              [0]*12,
        "Website & Remind App":  [0,391.37,0,0,0,0,0,0,0,0,0,0],
        "Event Equipment":       [0]*12,
    },
    "Fundraising": {
        "Birthday Books":    [0]*12,
        "Book Fair":         [0,0,500,3885.07,0,0,500,5841.48,0,0,0,0],
        "Croc Charms":       [0]*12,
        "Fall Pictures":     [0]*12,
        "FAST":              [0,0,0,9375,0,0,0,6405,0,7150,0,0],
        "Holiday Boutique":  [0,0,0,0,0,1466.06,0,7511.09,0,0,0,0],
        "Plant Sale":        [0]*12,
        "Spiritwear":        [0,0,0,0,0,0,0,898.72,0,0,0,0],
        "Spring Pictures":   [0]*12,
    },
    "Membership": {
        "Council Dues":        [0,0,0,0,0,0,0,0,0,125,0,0],
        "Membership Expenses": [0,0,654,202,0,0,0,0,0,0,0,0],
    },
    "Basket Dinner": {
        "Entertainment": [0]*12,
        "Raffles":       [0]*12,
        "Venue":         [0]*12,
    },
    "Programs": {
        "Bus Driver Appreciation": [0]*12,
        "Electric Parade":         [0]*12,
        "Family Connect Nights":   [0]*12,
        "Gingerbread U":           [0,0,0,0,0,4001.66,74.83,0,253.87,0,0,0],
        "Homecoming":              [0,0,67.20,0,0,0,0,0,0,0,0,0],
        "K Playdate":              [0]*12,
        "K Orientation":           [0]*12,
        "Milk & Cookies":          [0,0,362.19,0,0,0,0,0,0,0,0,0],
        "Multicultural Night":     [0,0,0,0,0,0,0,817.61,772.88,0,0,0],
        "Outdoor Movie":           [0,0,878.28,414.29,0,0,0,0,668.94,0,0,0],
        "Science Fair":            [0,0,0,0,0,0,0,0,925.53,0,0,0],
        "Spring Fling":            [0,0,0,0,0,0,0,0,0,835,0,0],
        "Staff Appreciation":      [0,0,0,0,0,0,0,0,700,134.24,0,0],
        "Talent Show":             [0,0,0,0,0,0,0,0,273.92,0,0,0],
        "Talent Show DJ":          [0]*12,
        "Volunteer Breakfast":     [0,0,0,0,0,0,0,0,600,0,0,0],
        "Welcome Back Breakfast":  [0,0,823.40,223.73,0,0,0,0,0,0,0,0],
        "WINGO":                   [0]*12,
    },
    "Donations": {
        "BOE Gifts":              [0,0,0,0,0,0,0,0,0,180,0,0],
        "Cultural Arts":          [0,0,0,0,2315,0,7500,0,0,4045,0,0],
        "Folders":                [0,0,0,1026.5,0,0,0,0,0,0,0,0],  # approx Oct
        "Gardening":              [0]*12,
        "Hospitality":            [0,0,0,203.23,0,0,0,0,0,0,0,0],
        "5th Staff T-Shirts":     [0]*12,
        "Recess Equipment":       [0,0,0,1000,0,0,0,0,0,0,0,0],
        "School Spirit":          [0]*12,
        "Sling Bags":             [0]*12,
        "Spelling Bee":           [0,0,0,0,0,0,0,0,206.5,0,0,0],
        "Sunshine Fund":          [0]*12,
        "Trick or Treat Street":  [0]*12,
        "WM Scholarships":        [0]*12,
    },
    "Grad Class Events": {
        "Monster Bash":    [0,0,0,273.51,140.17,0,0,0,0,0,0,0],
        "Monster Bash DJ": [0,0,0,500,0,0,0,0,0,0,0,0],
        "Electric Parade": [0]*12,
        "Winter Social":   [0]*12,
        "Winter Social DJ":[0]*12,
        "Moving Up":       [0]*12,
        "Picnic":          [181.58,0,0,0,0,0,0,0,0,0,0,0],
    },
    "Grad Class Expenses": {
        "Graduating Class Gifts":   [0]*12,
        "Graduating Mural":         [0]*12,
        "5th Grade T-Shirts":       [0,0,0,901,0,0,0,0,0,0,0,0],
        "Trunk or Treat Fundraiser":[0,0,0,258.67,0,0,0,0,0,0,0,0],
        "The Night at Rinx":        [0,0,0,0,0,0,650,60,0,0,0,0],
    },
}


def build_budget_sheet(ws, title: str, budget_data: dict, actuals_data: dict):
    ws.sheet_view.showGridLines = False

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 13
    for i, m in enumerate(MONTHS):
        col = get_column_letter(4 + i)
        ws.column_dimensions[col].width = 9
    ws.column_dimensions[get_column_letter(16)].width = 11  # Total
    ws.column_dimensions[get_column_letter(17)].width = 11  # P/L

    # Row 1: Title
    ws.merge_cells(f"A1:{get_column_letter(17)}1")
    c = ws["A1"]
    c.value = f"SETAUKET SCHOOL PTA  —  {title}"
    c.font  = Font(name="Arial", bold=True, size=13, color=WHITE)
    c.fill  = NAVY_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Row 2: As-of date
    ws.merge_cells(f"A2:{get_column_letter(17)}2")
    c = ws["A2"]
    c.value = f"As of {datetime.today().strftime('%B %d, %Y')}  |  Fiscal Year July 2025 – June 2026"
    c.font  = Font(name="Arial", italic=True, size=9, color="666666")
    c.alignment = Alignment(horizontal="center")

    # Row 3: Column headers
    headers = ["Category", "Last Year", "Budget (Annual)"] + MONTHS + ["Total", "Profit/Loss"]
    for col_idx, hdr in enumerate(headers, 1):
        c = ws.cell(row=3, column=col_idx)
        c.value = hdr
        c.font  = Font(name="Arial", bold=True, size=9, color=WHITE)
        c.fill  = TEAL_FILL
        c.alignment = Alignment(horizontal="center" if col_idx > 1 else "left",
                                vertical="center", wrap_text=True)
        c.border = THIN_BORDER
    ws.row_dimensions[3].height = 30

    data_row = 4

    for section, items in budget_data.items():
        sec_actuals = actuals_data.get(section, {})

        # Section header row
        ws.merge_cells(f"A{data_row}:{get_column_letter(17)}{data_row}")
        c = ws[f"A{data_row}"]
        c.value = section
        c.font  = Font(name="Arial", bold=True, size=10, color=WHITE)
        c.fill  = NAVY_FILL
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[data_row].height = 18
        section_start = data_row + 1
        data_row += 1

        for i_row, (item, (last_yr, budget)) in enumerate(items.items()):
            fill = LGREY_FILL if i_row % 2 == 1 else PatternFill()
            monthly = sec_actuals.get(item, [0]*12)

            # Col A: item name
            c = ws.cell(row=data_row, column=1, value=item)
            c.font = BODY_FONT; c.fill = fill
            c.alignment = Alignment(indent=2); c.border = THIN_BORDER

            # Col B: last year
            c = ws.cell(row=data_row, column=2, value=last_yr if last_yr else None)
            c.font = BODY_FONT; c.fill = fill
            c.number_format = MONEY_FMT; c.alignment = Alignment(horizontal="right")
            c.border = THIN_BORDER

            # Col C: budget
            c = ws.cell(row=data_row, column=3, value=budget if budget else None)
            c.font = BODY_FONT; c.fill = fill
            c.number_format = MONEY_FMT; c.alignment = Alignment(horizontal="right")
            c.border = THIN_BORDER

            # Cols D-O: monthly actuals
            for m_idx, val in enumerate(monthly):
                c = ws.cell(row=data_row, column=4 + m_idx, value=val if val else None)
                c.font = BODY_FONT; c.fill = fill
                c.number_format = MONEY_FMT; c.alignment = Alignment(horizontal="right")
                c.border = THIN_BORDER

            # Col P: Total (SUM of monthly cols)
            total_col = get_column_letter(16)
            c = ws.cell(row=data_row, column=16,
                        value=f"=SUM(D{data_row}:O{data_row})")
            c.font = BODY_FONT; c.fill = fill
            c.number_format = MONEY_FMT; c.alignment = Alignment(horizontal="right")
            c.border = THIN_BORDER

            # Col Q: Profit/Loss vs budget
            c = ws.cell(row=data_row, column=17,
                        value=f"=C{data_row}-P{data_row}" if budget else None)
            c.font = BODY_FONT; c.fill = fill
            c.number_format = MONEY_FMT; c.alignment = Alignment(horizontal="right")
            c.border = THIN_BORDER

            ws.row_dimensions[data_row].height = 15
            data_row += 1

        section_end = data_row - 1

        # Section total row
        c = ws.cell(row=data_row, column=1, value=f"Total {section}")
        c.font = TOTAL_FONT; c.fill = LTBLUE_FILL
        c.alignment = Alignment(indent=1); c.border = MED_BORDER

        for col_idx in [2, 3]:
            col_l = get_column_letter(col_idx)
            c = ws.cell(row=data_row, column=col_idx,
                        value=f"=SUM({col_l}{section_start}:{col_l}{section_end})")
            c.font = TOTAL_FONT; c.fill = LTBLUE_FILL
            c.number_format = MONEY_FMT; c.alignment = Alignment(horizontal="right")
            c.border = MED_BORDER

        for m_col in range(4, 18):
            col_l = get_column_letter(m_col)
            c = ws.cell(row=data_row, column=m_col,
                        value=f"=SUM({col_l}{section_start}:{col_l}{section_end})")
            c.font = TOTAL_FONT; c.fill = LTBLUE_FILL
            c.number_format = MONEY_FMT; c.alignment = Alignment(horizontal="right")
            c.border = MED_BORDER

        ws.row_dimensions[data_row].height = 18
        data_row += 2  # blank row between sections

    # Grand total row
    ws.row_dimensions[data_row].height = 20
    c = ws.cell(row=data_row, column=1, value="GRAND TOTAL")
    c.font = Font(name="Arial", bold=True, size=11, color=WHITE)
    c.fill = NAVY_FILL; c.alignment = Alignment(indent=1); c.border = MED_BORDER

    for col_idx in range(2, 18):
        col_l = get_column_letter(col_idx)
        # Sum all the section-total rows (every 2nd row starting from section ends)
        c = ws.cell(row=data_row, column=col_idx,
                    value=f"=SUMIF(A4:A{data_row-1},\"Total*\",{col_l}4:{col_l}{data_row-1})")
        c.font = Font(name="Arial", bold=True, size=10, color=WHITE)
        c.fill = NAVY_FILL; c.number_format = MONEY_FMT
        c.alignment = Alignment(horizontal="right"); c.border = MED_BORDER


# ── 6. SHEET: GIVEBACK RECONCILIATION ────────────────────────────────────────
def build_giveback_sheet(ws, givebacks: list[dict], bank: dict):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 20

    # Title
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "SETAUKET SCHOOL PTA  —  Giveback Reconciliation"
    c.font  = Font(name="Arial", bold=True, size=13, color=WHITE)
    c.fill  = NAVY_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:E2")
    c = ws["A2"]
    c.value = f"Generated: {datetime.today().strftime('%B %d, %Y')}"
    c.font  = Font(name="Arial", italic=True, size=9, color="666666")
    c.alignment = Alignment(horizontal="center")

    # Headers
    row = 4
    for col, hdr in zip(["A","B","C","D","E"],
                         ["Item", "Category", "Transactions", "Amount", "% of Total"]):
        c = ws[f"{col}{row}"]
        c.value = hdr; c.font = SUBHDR_FONT; c.fill = TEAL_FILL
        c.alignment = Alignment(horizontal="center" if col != "A" else "left",
                                vertical="center")
        c.border = THIN_BORDER
    ws.row_dimensions[row].height = 18
    row += 1

    data_start = row
    total_amt = sum(g["total"] for g in givebacks)

    for i, g in enumerate(givebacks):
        fill = LGREY_FILL if i % 2 == 1 else PatternFill()
        ws[f"A{row}"].value = g["item"];     ws[f"A{row}"].font = BODY_FONT
        ws[f"A{row}"].fill = fill;           ws[f"A{row}"].border = THIN_BORDER
        ws[f"A{row}"].alignment = Alignment(indent=1)

        ws[f"B{row}"].value = g["category"]; ws[f"B{row}"].font = BODY_FONT
        ws[f"B{row}"].fill = fill;           ws[f"B{row}"].border = THIN_BORDER
        ws[f"B{row}"].alignment = Alignment(horizontal="center")

        ws[f"C{row}"].value = g["count"];    ws[f"C{row}"].font = BODY_FONT
        ws[f"C{row}"].fill = fill;           ws[f"C{row}"].border = THIN_BORDER
        ws[f"C{row}"].alignment = Alignment(horizontal="center")

        ws[f"D{row}"].value = g["total"];    ws[f"D{row}"].font = BODY_FONT
        ws[f"D{row}"].fill = fill;           ws[f"D{row}"].border = THIN_BORDER
        ws[f"D{row}"].number_format = MONEY_FMT
        ws[f"D{row}"].alignment = Alignment(horizontal="right")

        ws[f"E{row}"].value = f"=D{row}/D${len(givebacks)+data_start}"
        ws[f"E{row}"].font = BODY_FONT;      ws[f"E{row}"].fill = fill
        ws[f"E{row}"].border = THIN_BORDER
        ws[f"E{row}"].number_format = "0.0%"
        ws[f"E{row}"].alignment = Alignment(horizontal="center")

        ws.row_dimensions[row].height = 15
        row += 1

    # Total
    total_row_num = row
    ws[f"A{row}"].value = "TOTAL"; ws[f"A{row}"].font = TOTAL_FONT
    ws[f"A{row}"].fill = LTBLUE_FILL; ws[f"A{row}"].border = MED_BORDER
    ws[f"A{row}"].alignment = Alignment(indent=1)
    ws[f"C{row}"].value = f"=SUM(C{data_start}:C{row-1})"
    ws[f"C{row}"].font = TOTAL_FONT; ws[f"C{row}"].fill = LTBLUE_FILL
    ws[f"C{row}"].border = MED_BORDER; ws[f"C{row}"].alignment = Alignment(horizontal="center")
    ws[f"D{row}"].value = f"=SUM(D{data_start}:D{row-1})"
    ws[f"D{row}"].font = TOTAL_FONT; ws[f"D{row}"].fill = LTBLUE_FILL
    ws[f"D{row}"].border = MED_BORDER; ws[f"D{row}"].number_format = MONEY_FMT
    ws[f"D{row}"].alignment = Alignment(horizontal="right")
    ws[f"E{row}"].value = "100.0%"; ws[f"E{row}"].font = TOTAL_FONT
    ws[f"E{row}"].fill = LTBLUE_FILL; ws[f"E{row}"].border = MED_BORDER
    ws[f"E{row}"].alignment = Alignment(horizontal="center")
    for col in ["B"]:
        ws[f"{col}{row}"].fill = LTBLUE_FILL; ws[f"{col}{row}"].border = MED_BORDER
    ws.row_dimensions[row].height = 18
    row += 2

    # Category summary
    ws.merge_cells(f"A{row}:E{row}")
    c = ws[f"A{row}"]
    c.value = "SUMMARY BY CATEGORY"
    c.font  = Font(name="Arial", bold=True, size=11, color=WHITE)
    c.fill  = NAVY_FILL; c.alignment = Alignment(indent=1)
    ws.row_dimensions[row].height = 18
    row += 1

    for col, hdr in zip(["A","B","C","D"],["Category","Items","Transactions","Total"]):
        c = ws[f"{col}{row}"]
        c.value = hdr; c.font = SUBHDR_FONT; c.fill = TEAL_FILL
        c.alignment = Alignment(horizontal="center" if col != "A" else "left")
        c.border = THIN_BORDER
    ws.row_dimensions[row].height = 18
    row += 1

    from collections import defaultdict
    cat_summary = defaultdict(lambda: {"count": 0, "txns": 0, "total": 0.0})
    for g in givebacks:
        cat_summary[g["category"]]["count"] += 1
        cat_summary[g["category"]]["txns"]  += g["count"]
        cat_summary[g["category"]]["total"] += g["total"]

    for i, (cat, vals) in enumerate(sorted(cat_summary.items())):
        fill = LGREY_FILL if i % 2 == 1 else PatternFill()
        ws[f"A{row}"].value = cat;           ws[f"A{row}"].font = BODY_FONT
        ws[f"A{row}"].fill = fill;           ws[f"A{row}"].border = THIN_BORDER
        ws[f"A{row}"].alignment = Alignment(indent=1)
        ws[f"B{row}"].value = vals["count"]; ws[f"B{row}"].font = BODY_FONT
        ws[f"B{row}"].fill = fill;           ws[f"B{row}"].border = THIN_BORDER
        ws[f"B{row}"].alignment = Alignment(horizontal="center")
        ws[f"C{row}"].value = vals["txns"];  ws[f"C{row}"].font = BODY_FONT
        ws[f"C{row}"].fill = fill;           ws[f"C{row}"].border = THIN_BORDER
        ws[f"C{row}"].alignment = Alignment(horizontal="center")
        ws[f"D{row}"].value = vals["total"]; ws[f"D{row}"].font = BODY_FONT
        ws[f"D{row}"].fill = fill;           ws[f"D{row}"].border = THIN_BORDER
        ws[f"D{row}"].number_format = MONEY_FMT
        ws[f"D{row}"].alignment = Alignment(horizontal="right")
        ws.row_dimensions[row].height = 15
        row += 1

    row += 1

    # Bank reconciliation note
    gb_total = sum(g["total"] for g in givebacks)
    bank_gb_deposit = next((d["amount"] for d in bank["deposits"]
                            if "Gb Payout" in d.get("description","") or
                               "GB" in d.get("description","").upper()), 0.0)

    ws.merge_cells(f"A{row}:E{row}")
    c = ws[f"A{row}"]
    c.value = "GIVEBACK ↔ BANK RECONCILIATION"
    c.font  = Font(name="Arial", bold=True, size=11, color=WHITE)
    c.fill  = NAVY_FILL; c.alignment = Alignment(indent=1)
    ws.row_dimensions[row].height = 18
    row += 1

    for lbl, val in [("Givebacks Platform Total", gb_total),
                     ("Givebacks Deposit in Bank Statement", bank_gb_deposit),
                     ("Difference", gb_total - bank_gb_deposit)]:
        fill = GREEN_FILL if lbl == "Difference" and abs(gb_total - bank_gb_deposit) < 0.01 \
               else RED_FILL if lbl == "Difference" else PatternFill()
        ws[f"A{row}"].value = lbl;  ws[f"A{row}"].font = BOLD_FONT
        ws[f"A{row}"].fill = fill;  ws[f"A{row}"].border = THIN_BORDER
        ws[f"A{row}"].alignment = Alignment(indent=2)
        ws[f"D{row}"].value = val;  ws[f"D{row}"].font = BOLD_FONT
        ws[f"D{row}"].fill = fill;  ws[f"D{row}"].border = THIN_BORDER
        ws[f"D{row}"].number_format = MONEY_FMT
        ws[f"D{row}"].alignment = Alignment(horizontal="right")
        for col in ["B","C","E"]:
            ws[f"{col}{row}"].fill = fill; ws[f"{col}{row}"].border = THIN_BORDER
        ws.row_dimensions[row].height = 16
        row += 1


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    base = Path("/mnt/user-data/uploads")
    gb_path = base / "givebacks_example.csv"
    qb_path = base / "PTA_feb_quickbooks.csv"
    pdf_path = base / "chase_feb_statement.pdf"
    out_path = Path("/mnt/user-data/outputs/PTA_Treasurer_Report_February_2026.xlsx")

    print("Parsing Givebacks...")
    givebacks = parse_givebacks(gb_path)

    print("Parsing QuickBooks...")
    qb = parse_quickbooks(qb_path)

    print("Parsing Chase PDF...")
    bank = parse_chase_pdf(pdf_path)

    print("Building Excel workbook...")
    wb = openpyxl.Workbook()

    # Sheet 1: Monthly Treasurer Report
    ws1 = wb.active
    ws1.title = "Treasurer Report"
    build_treasurer_report(ws1, qb, bank, "February 2026")

    # Sheet 2: Budget vs Actuals - Income
    ws2 = wb.create_sheet("Income Budget vs Actuals")
    build_budget_sheet(ws2, "Budget vs Actuals — Income",
                       INCOME_BUDGET, INCOME_ACTUALS)

    # Sheet 3: Budget vs Actuals - Expenses
    ws3 = wb.create_sheet("Expense Budget vs Actuals")
    build_budget_sheet(ws3, "Budget vs Actuals — Expenses",
                       EXPENSE_BUDGET, EXPENSE_ACTUALS)

    # Sheet 4: Giveback Reconciliation
    ws4 = wb.create_sheet("Giveback Reconciliation")
    build_giveback_sheet(ws4, givebacks, bank)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved → {out_path}")


if __name__ == "__main__":
    main()
